from openpyxl import load_workbook
from ubicacionv3 import cargar_horarios, get_ubicaciones_exportados, get_cajeros
from getSelfxDia import get_self_exportados
from ExcelUbicacionv1 import asignar_cajeros, asignar_selfs
from InhabilitadosToggle import get_inhabilitados_indices

if (__name__ == '__main__'):
    # Cargar el archivo Excel y la hoja de trabajo
    archivo_origen = 'Plantilla.xlsx'
    archivo_destino = 'PlantillaSemanal_Exportada.xlsx'
    inhabilitados_path = "inhabilitados.txt"

    horarios = cargar_horarios("horario.txt")

    wb = load_workbook(archivo_origen)
    hoja_original = wb.active
    for i in range(1, 8):
        nueva_hoja = wb.copy_worksheet(hoja_original)
        nueva_hoja.title = str(i)

    cajeros = get_cajeros(horarios)
    inhabilitados_indices = get_inhabilitados_indices(cajeros, inhabilitados_path)
    
    for dia_seleccionado in range(1, 8):
        hoja = wb[f"{dia_seleccionado}"]
        cajeros = get_ubicaciones_exportados(horarios, dia_seleccionado, inhabilitados_indices)
        selfs = get_self_exportados(horarios, dia_seleccionado)
        asignar_cajeros(cajeros, hoja)
        asignar_selfs(selfs, hoja)

    del wb[wb.sheetnames[0]]

    # Guardar el nuevo archivo
    wb.save(archivo_destino)

    print("Asignación completada y archivo guardado como", archivo_destino)
