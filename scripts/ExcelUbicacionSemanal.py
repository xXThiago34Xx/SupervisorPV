from openpyxl import load_workbook
from ubicacionv3 import cargar_horarios, dia_menu, inhabilitados_menu, get_ubicaciones_exportados
from getSelfxDia import get_self_exportados
from ExcelUbicacionv1 import asignar_cajeros, asignar_selfs

if (__name__ == '__main__'):
    # Cargar el archivo Excel y la hoja de trabajo
    archivo_origen = 'Plantilla.xlsx'
    archivo_destino = 'PlantillaSemanal_Exportada.xlsx'
    
    horarios = cargar_horarios("horario.txt")

    wb = load_workbook(archivo_origen)
    hoja_original = wb.active
    for i in range(1, 8):
        nueva_hoja = wb.copy_worksheet(hoja_original)
        nueva_hoja.title = str(i)

    inhabilitados_indices = inhabilitados_menu(horarios)
    
    for dia_seleccionado in range(1, 8):
        hoja = wb[f"{dia_seleccionado}"]
        cajeros = get_ubicaciones_exportados(horarios, dia_seleccionado, inhabilitados_indices)
        selfs = get_self_exportados(horarios, dia_seleccionado)
        asignar_cajeros(cajeros, hoja)
        asignar_selfs(selfs, hoja)

    del wb[wb.sheetnames[0]]

    # Guardar el nuevo archivo
    wb.save(archivo_destino)

    print("Asignación completada y archivo guardado como", archivo_destino)
