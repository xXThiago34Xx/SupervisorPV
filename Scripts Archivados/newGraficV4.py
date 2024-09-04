import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from collections import defaultdict
import os

# Diccionario para asociar colores con tipos de personal
COLORES_PERSONAL = {
    "Asistente de Self Checkout": "FFFF00",    # Amarillo
    "Representante de Servicio": "ADD8E6",     # Azul Claro
    "Cajer@": "FF0000",                        # Rojo
    "Ecommerce": "EE82EE",                      # Violeta
    "Supervisor(@)": "90EE90"                   # Verde Claro
}

def cargar_horarios(archivo):
    horarios = []
    with open(archivo, "r") as archivo:
        lector = csv.reader(archivo)
        for fila in lector:
            horarios.append(fila)
    return horarios

def formatear_hora(hora):
    hora_str = str(hora)
    if len(hora_str) == 1 or len(hora_str) == 2:
        hora_str = hora_str.zfill(2) + ":00"
    elif len(hora_str) == 3:
        hora_str = "0" + hora_str[0] + ":" + hora_str[1:]
    elif len(hora_str) == 4:
        hora_str = hora_str[:2] + ":" + hora_str[2:]
    return hora_str

def obtener_horarios_por_dia(horarios, dia, tipo_personal):
    dia_index = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"].index(dia.lower())
    horarios_dia = defaultdict(list)
    
    for registro in horarios:
        apellidos, nombres, cargo, *dias = registro
        if cargo.lower() == tipo_personal.lower():
            entrada = dias[dia_index * 2]
            salida = dias[dia_index * 2 + 1]
            entrada_formateada = formatear_hora(entrada)
            salida_formateada = formatear_hora(salida)
            if entrada_formateada != "DESCANSO" and salida_formateada != "DESCANSO":
                horarios_dia[(nombres, apellidos)] = (entrada_formateada, salida_formateada)
    
    # Ordenar por hora de entrada
    horarios_dia_ordenados = dict(sorted(horarios_dia.items(), key=lambda x: x[1][0]))
    
    return horarios_dia_ordenados

def generar_intervalos():
    intervalos = []
    for hora in range(6, 24):
        for minuto in [0, 15, 30, 45]:
            inicio = f"{str(hora).zfill(2)}:{str(minuto).zfill(2)}"
            if minuto == 45:
                fin = f"{str(hora + 1).zfill(2)}:00"
            else:
                fin = f"{str(hora).zfill(2)}:{str(minuto + 15).zfill(2)}"
            intervalos.append(f"{inicio} - {fin}")
    return intervalos

def crear_hoja(ws, intervalos, horarios_dia, tipo_personal):
    encabezados = ["Intervalos"] + [f"{nombres} {apellidos}" for nombres, apellidos in horarios_dia.keys()] + ["#"]
    ws.append(encabezados)
    
    for i, intervalo in enumerate(intervalos, start=2):
        ws[f"A{i}"] = intervalo
    
    for col, (nombres, apellidos) in enumerate(horarios_dia.keys(), start=2):
        entrada, salida = horarios_dia[(nombres, apellidos)]
        for i, intervalo in enumerate(intervalos, start=2):
            inicio_intervalo = intervalo.split(" - ")[0]
            if entrada <= inicio_intervalo < salida:
                ws.cell(row=i, column=col).value = "EN TURNO"
                ws.cell(row=i, column=col).fill = PatternFill(
                    start_color=COLORES_PERSONAL[tipo_personal], 
                    end_color=COLORES_PERSONAL[tipo_personal], 
                    fill_type="solid"
                )
    
    for i, intervalo in enumerate(intervalos, start=2):
        count = sum(1 for col in range(2, len(horarios_dia) + 2) if ws.cell(row=i, column=col).value == "EN TURNO")
        ws.cell(row=i, column=len(horarios_dia) + 2).value = count

def generar_excel(horarios, dia, tipo_personal):
    intervalos = generar_intervalos()
    wb = openpyxl.Workbook()
    
    if tipo_personal == "Todos":
        tipos_personal = ["Asistente de Self Checkout", "Representante de Servicio", "Cajer@", "Ecommerce", "Supervisor(@)"]
        for tipo in tipos_personal:
            ws = wb.create_sheet(title=tipo)
            horarios_dia = obtener_horarios_por_dia(horarios, dia, tipo)
            crear_hoja(ws, intervalos, horarios_dia, tipo)
    else:
        ws = wb.active
        ws.title = tipo_personal
        horarios_dia = obtener_horarios_por_dia(horarios, dia, tipo_personal)
        crear_hoja(ws, intervalos, horarios_dia, tipo_personal)
    
    archivo_salida = f"Horario_{tipo_personal}_{dia}.xlsx"
    wb.save(archivo_salida)
    print(f"Archivo {archivo_salida} generado exitosamente.")

def main():
    archivo_horarios = "horario.txt"
    horarios = cargar_horarios(archivo_horarios)
    
    print("\nSeleccionar Día de la Semana:")
    dias_semana = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    for i, dia in enumerate(dias_semana, 1):
        print(f"{i}. {dia.capitalize()}")
    dia_opcion = int(input("Ingrese el número de opción: ")) - 1
    dia = dias_semana[dia_opcion]
    
    print("\nSeleccionar Tipo de Personal:")
    tipos_personal = ["Asistente de Self Checkout", "Representante de Servicio", "Cajer@", "Ecommerce", "Supervisor(@)", "Todos"]
    for i, tipo in enumerate(tipos_personal, 1):
        print(f"{i}. {tipo}")
    tipo_opcion = int(input("Ingrese el número de opción: ")) - 1
    tipo_personal = tipos_personal[tipo_opcion]
    
    generar_excel(horarios, dia, tipo_personal)

if __name__ == "__main__":
    main()