import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import datetime
import csv
from collections import defaultdict
from mostrarNicks import getNick, getApNick
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# Definición de los colores por roles
roles_colores = {
    "Self Checkout": "FFFF00",  # Amarillo
    "RS": "ADD8E6",             # Azul Claro
    "Cajer@": "FF0000",         # Rojo
    "Ecommerce": "EE82EE",      # Violeta
    "Supervisor(@)": "90EE90"   # Verde Claro
}

# Función para obtener la hora correspondiente a un número entero en intervalos de 15 minutos
def getHoraxIntervalo(numero):
    """
    Devuelve la hora en formato HH:MM correspondiente al número entero dado,
    basado en intervalos de 15 minutos desde las 6:00 hasta las 24:00.
    
    :param numero: int, el número entero que representa el intervalo.
    :return: str, la hora en formato HH:MM.
    """
    if numero < 1 or numero > 73:
        raise ValueError("El número debe estar entre 1 y 73.")
    
    base_hora = 6
    intervalos_por_hora = 4  # 4 intervalos de 15 minutos por hora
    minutos_por_intervalo = 15
    
    intervalo = numero - 1
    hora = base_hora + intervalo // intervalos_por_hora
    minutos = (intervalo % intervalos_por_hora) * minutos_por_intervalo
    
    return f"{hora:02}:{minutos:02}"

# Función para llenar las entradas en el Excel
def llenar_entradas(ws, Entradas):
    """
    Escribe los datos del arreglo 'Entradas' en las columnas correspondientes de la hoja de trabajo (ws).
    
    :param ws: La hoja de trabajo donde se escribirán los datos de Entradas.
    :param Entradas: Matriz con las entradas a escribir.
    """
    for i in range(1, 73):  # Recorrer las horas de entrada (1 a 72)
        hora = getHoraxIntervalo(i)  # Obtener la hora en formato HH:MM
        
        # Colocar la hora en la columna A
        celda_hora = f"A{i+2}"  # La columna A, desde la fila 3
        ws[celda_hora] = hora

        # Recorrer los elementos de Entradas[i] y colocarlos en las columnas B a G
        for j, entrada in enumerate(Entradas[i]):
            if j < 6 and len(entrada) == 2:  # Solo colocar en las primeras 6 columnas (B a G) si hay 2 elementos (nombre y tipo)
                nombre, tipo_personal = entrada
                
                # Definir la celda correspondiente en la hoja
                celda_columna = get_column_letter(j + 2)  # Desde la columna B (que es la columna 2)
                celda_fila = i + 2  # Fila empieza en 3 porque fila 1 y 2 son los encabezados
                celda = f"{celda_columna}{celda_fila}"
                
                # Escribir el nombre en la celda
                ws[celda] = nombre
                
                # Aplicar color según el tipo de personal
                if tipo_personal in roles_colores:
                    fill = PatternFill(start_color=roles_colores[tipo_personal], end_color=roles_colores[tipo_personal], fill_type='solid')
                    ws[celda].fill = fill
                
                # Aplicar bordes a la celda
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws[celda].border = border

def llenar_salidas(ws, Salidas):
    """
    Escribe los datos del arreglo 'Entradas' en las columnas correspondientes de la hoja de trabajo (ws).
    
    :param ws: La hoja de trabajo donde se escribirán los datos de Entradas.
    :param Entradas: Matriz con las entradas a escribir.
    """
    for i in range(1, 73):  # Recorrer las horas de entrada (1 a 72)
        hora = getHoraxIntervalo(i)  # Obtener la hora en formato HH:MM
        
        # Colocar la hora en la columna A
        celda_hora = f"H{i+2}"  # La columna H, desde la fila 3
        ws[celda_hora] = hora

        # Recorrer los elementos de Entradas[i] y colocarlos en las columnas B a G
        for j, salida in enumerate(Salidas[i]):
            if j < 7 and len(salida) == 2:  # Solo colocar en las primeras 6 columnas (B a G) si hay 2 elementos (nombre y tipo)
                nombre, tipo_personal = salida
                
                # Definir la celda correspondiente en la hoja
                celda_columna = get_column_letter(j + 2 + 7)  # Desde la columna B (que es la columna 2)
                celda_fila = i + 2  # Fila empieza en 3 porque fila 1 y 2 son los encabezados
                celda = f"{celda_columna}{celda_fila}"
                
                # Escribir el nombre en la celda
                ws[celda] = nombre
                
                # Aplicar color según el tipo de personal
                if tipo_personal in roles_colores:
                    fill = PatternFill(start_color=roles_colores[tipo_personal], end_color=roles_colores[tipo_personal], fill_type='solid')
                    ws[celda].fill = fill
                
                # Aplicar bordes a la celda
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws[celda].border = border

def rellenar(ws, celda, texto='', color_rgb=''):
    """
    Escribe un texto en una celda y cambia el color de fondo de esa celda si se proporcionan valores no vacíos.
    
    :param ws: La hoja de trabajo donde se escribirá el texto.
    :param celda: La referencia de la celda (ejemplo: 'A1').
    :param texto: El texto que se escribirá en la celda. Si está vacío, no se modificará el texto.
    :param color_rgb: El color de fondo en formato RGB (ejemplo: 'FFFF00' para amarillo). Si está vacío, no se modificará el color.
    """
    # Escribir el texto en la celda si no está vacío
    if texto:
        ws[celda] = texto
    
    # Cambiar el color de fondo de la celda si se proporciona un color
    if color_rgb:
        fill = PatternFill(start_color=color_rgb, end_color=color_rgb, fill_type='solid')
        ws[celda].fill = fill

def combinar_y_centrar_celdas(ws, celda1, celda2, texto):
    """
    Combina dos celdas y centra el texto en el rango combinado.
    
    :param ws: La hoja de trabajo donde se realizará la combinación y el centrado.
    :param celda1: La referencia de la primera celda (ejemplo: 'A1').
    :param celda2: La referencia de la segunda celda (ejemplo: 'B2').
    :param texto: El texto que se escribirá en el rango combinado.
    """
    # Obtener el rango de celdas
    rango = ws[celda1:celda2]
    
    # Escribir el texto en la celda superior izquierda del rango combinado
    ws[celda1] = texto
    
    # Combinar las celdas
    ws.merge_cells(start_row=rango[0][0].row, start_column=rango[0][0].column,
                   end_row=rango[-1][-1].row, end_column=rango[-1][-1].column)
    
    # Centrar el texto
    ws[celda1].alignment = Alignment(horizontal='center', vertical='center')


# Función para crear el archivo Excel con el formato de la tabla de horarios
def crearHojaXDia(wb, Entradas, Salidas, dia):
    """
    Crea un archivo Excel con el formato de tabla para los horarios del personal,
    llenando los datos desde la matriz 'Entradas'.
    
    :param Entradas: Matriz con las entradas del personal.
    """
    # Crear el libro y la hoja de trabajo
    
    wb.create_sheet(f"{dia}")
    ws = wb[f"{dia}"]

    #Izquierda
    rellenar(ws, "B1", "Cajer@s", roles_colores["Cajer@"])
    rellenar(ws, "C1", "RS", roles_colores["RS"])
    rellenar(ws, "D1", "Self Checkout", roles_colores["Self Checkout"])
    rellenar(ws, "E1", "Ecommerce", roles_colores["Ecommerce"])
    rellenar(ws, "F1", "Supervisores", roles_colores["Supervisor(@)"])

    #Derecha
    rellenar(ws, "J1", "Cajer@s", roles_colores["Cajer@"])
    rellenar(ws, "K1", "RS", roles_colores["RS"])
    rellenar(ws, "L1", "Self Checkout", roles_colores["Self Checkout"])
    rellenar(ws, "M1", "Ecommerce", roles_colores["Ecommerce"])
    rellenar(ws, "N1", "Supervisores", roles_colores["Supervisor(@)"])

    #Titulos
    rellenar(ws, "A1", "Entradas", "")
    rellenar(ws, "H1", "Salidas", "")
    combinar_y_centrar_celdas(ws, "A2", "G2", "Horas de Entrada del Personal del Area de Cajas")
    combinar_y_centrar_celdas(ws, "H2", "O2", "Horas de Salida del Personal del Area de Cajas")



    # Definir los colores para las filas
    color_gris = 'D3D3D3'  # Gris claro
    color_blanco = 'FFFFFF'  # Blanco

    # Definir los bordes
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    # Aplicar bordes a todas las celdas desde la columna A hasta la columna O
    for row in ws.iter_rows(min_row=1, max_row=74, min_col=1, max_col=15):  # Ajusta el rango según sea necesario
        for cell in row:
            cell.border = border
            
        # Crear una lista de horarios cada 15 minutos desde las 6:00 hasta las 24:00
        start_time = datetime.strptime("06:00", "%H:%M")
        end_time = datetime.strptime("23:45", "%H:%M")
        current_time = start_time

        horarios = []
        while current_time <= end_time:
            horarios.append(current_time.strftime("%H:%M"))
            current_time += timedelta(minutes=15)

    # Rellenar las columnas A y H con los horarios
    for i, horario in enumerate(horarios):
        row_num = 3 + i
        if row_num <= 75:  # Asegurarse de no exceder la fila 75
            ws[f'A{row_num}'] = horario
            ws[f'H{row_num}'] = horario

            # Alternar colores de las filas
            fill_color = color_gris if (row_num - 3) % 2 == 0 else color_blanco
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            for col in range(1, 16):  # Desde la columna A hasta la columna O
                col_letter = get_column_letter(col)
                ws[f'{col_letter}{row_num}'].fill = fill

    # Ajustar el ancho de las columnas para una mejor visualización
    for col in range(1, 16):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 15
    
    # Llenar las entradas en la hoja
    llenar_entradas(ws, Entradas)
    llenar_salidas(ws, Salidas)



def cargar_horarios(archivo):
    horarios = []
    with open(archivo, "r") as archivo:
        lector = csv.reader(archivo)
        for fila in lector:
            horarios.append(fila)
    return horarios

def obtener_horarios_por_dia(horarios, dia, tipo_personal):
    dia_index = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"].index(dia.lower())
    horarios_dia = defaultdict(list)
    
    for registro in horarios:
        apellidos, nombres, cargo, *dias = registro
        if cargo.lower() == tipo_personal.lower():
            entrada = dias[dia_index * 2]
            salida = dias[dia_index * 2 + 1]
            entrada_formateada = formatear_hora(entrada)
            salida_formateada = formatear_hora(salida)
            if entrada_formateada == "DESCANSO" and salida_formateada == "DESCANSO":
                horarios_dia['DESCANSO'].append((apellidos, nombres))
            else:
                horarios_dia[entrada_formateada].append((apellidos, nombres, entrada_formateada, salida_formateada))
                horarios_dia[salida_formateada].append((apellidos, nombres, entrada_formateada, salida_formateada))
    
    return horarios_dia

def formatear_hora(hora):
    hora_str = str(hora)
    if len(hora_str) == 1 or len(hora_str) == 2:
        hora_str = hora_str.zfill(2) + ":00"
    elif len(hora_str) == 3:
        hora_str = "0" + hora_str[0] + ":" + hora_str[1:]
    elif len(hora_str) == 4:
        hora_str = hora_str[:2] + ":" + hora_str[2:]
    return hora_str

def ESaMEMORIA(horarios_dia, tipoPersonal):
    horarios_entradas = defaultdict(list)
    horarios_salidas = defaultdict(list)

    for hora, registros in horarios_dia.items():
        if hora == 'DESCANSO':
            continue
        for apellidos, nombres, entrada, salida in registros:
            if hora == entrada:
                horarios_entradas[hora].append((apellidos, nombres, entrada, salida))
            elif hora == salida:
                horarios_salidas[hora].append((apellidos, nombres, entrada, salida))

    for hora in sorted(set(horarios_entradas.keys()).union(horarios_salidas.keys())):
        print(f"\n===== {hora} =====")
        if hora in horarios_entradas:
            print("Entradas:")
            for apellidos, nombres, entrada, salida in sorted(horarios_entradas[hora], key=lambda x: x[1]):
                print(f"{apellidos} {nombres} - {entrada} - {salida}")
                nombre_completo = f"{nombres} {apellidos}"
                nick = getNick(nombre_completo)
                ap = getApNick(nombre_completo)
                nickname_completo = f"{nick}" # f"{nick} {ap}"
                Empilar(Entradas, getNHora(entrada), nickname_completo, tipoPersonal)
        if hora in horarios_salidas:
            print("Salidas:")
            for apellidos, nombres, entrada, salida in sorted(horarios_salidas[hora], key=lambda x: x[1]):
                print(f"{apellidos} {nombres} - {entrada} - {salida}")
                nombre_completo = f"{nombres} {apellidos}"
                nick = getNick(nombre_completo)
                ap = getApNick(nombre_completo)
                nickname_completo = f"{nick}" # f"{nick} {ap}"
                Empilar(Salidas, getNHora(salida), nickname_completo, tipoPersonal)

def Empilar(arreglo, fila, valor1, valor2):
    """
    Inserta dos valores en la siguiente posición vacía de la fila especificada en el arreglo tridimensional.
    
    :param arreglo: List[List[List[str]]], el arreglo tridimensional en el que se van a insertar los valores.
    :param fila: int, el índice de la fila donde se desea insertar los valores.
    :param valor1: str, el primer valor a insertar en la fila especificada.
    :param valor2: str, el segundo valor a insertar en la fila especificada.
    :return: None
    """
    # Validar que el índice de fila esté dentro del rango
    if fila < 0 or fila >= len(arreglo):
        raise IndexError("El índice de la fila está fuera del rango.")
    
    # Buscar la siguiente posición vacía en la fila especificada
    for columna in range(len(arreglo[fila])):
        if arreglo[fila][columna] == ["", ""]:
            arreglo[fila][columna] = [valor1, valor2]
            return
    
    # Si no se encuentra una posición vacía, levantar una excepción
    #raise Exception("No hay espacio disponible en la fila especificada.")

def getNHora(hora_str):
    # Convertir el string de hora a una lista con horas y minutos
    hora, minutos = map(int, hora_str.split(":"))
    
    # Validar que la hora esté entre 6:00 y 24:00
    if 6 <= hora <= 24:
        # Calcular el número asociado a la hora en formato de pasos de 15 minutos
        # Por cada hora que pase después de las 6, sumamos 4 (cada hora tiene 4 intervalos de 15 minutos)
        numero_hora = (hora - 6) * 4
        
        # Sumar el valor dependiendo de los minutos
        if minutos == 15:
            numero_hora += 1
        elif minutos == 30:
            numero_hora += 2
        elif minutos == 45:
            numero_hora += 3
        elif minutos != 0:
            raise ValueError("Los minutos deben ser 0, 15, 30 o 45")
        
        # El número final se incrementa en 1 ya que "6:00" corresponde a 1
        return numero_hora + 1
    else:
        # Si la hora está fuera del rango permitido
        raise ValueError("La hora debe estar entre 6:00 y 24:00")


def EliminarPrimeraHoja():
    # Cargar el archivo de Excel
    workbook = load_workbook("ES_Semanal.xlsx")

    # Obtener el nombre de la primera hoja
    first_sheet_name = workbook.sheetnames[0]

    # Eliminar la primera hoja
    del workbook[first_sheet_name]

    # Guardar el archivo modificado
    workbook.save("ES_Semanal.xlsx")
        







archivo_horarios = "horario.txt"
horarios = cargar_horarios(archivo_horarios)
dias = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
tipos_personal = ["Cajer@", "RS", "Self Checkout", "Ecommerce", "Supervisor(@)"]  


Entradas = [[["" for _ in range(2)] for _ in range(6)] for _ in range(73)]
Salidas = [[["" for _ in range(2)] for _ in range(7)] for _ in range(73)]
wb = openpyxl.Workbook()
for dia in dias:
    for tipoPersonal in tipos_personal:
        #tipo_personal = tipos_personal[tipoPersonal]
        horarios_dia = obtener_horarios_por_dia(horarios, dia, tipoPersonal)
        print(f"----- {tipoPersonal} -----")
        ESaMEMORIA(horarios_dia, tipoPersonal)
    crearHojaXDia(wb, Entradas, Salidas, dia)
    Entradas = [[["" for _ in range(2)] for _ in range(6)] for _ in range(73)]
    Salidas = [[["" for _ in range(2)] for _ in range(7)] for _ in range(73)]

# Guardar el archivo Excel
wb.save(f"ES_Semanal.xlsx")
EliminarPrimeraHoja()
print(f"Archivo 'ES_Semanal.xlsx' creado correctamente.")