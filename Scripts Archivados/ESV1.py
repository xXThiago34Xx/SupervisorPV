import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Definición de los colores por roles
roles_colores = {
    "Self Checkout": "FFFF00",  # Amarillo
    "RS": "ADD8E6",             # Azul Claro
    "Cajer@": "FF0000",         # Rojo
    "Ecommerce": "EE82EE",      # Violeta
    "Supervisor(@)": "90EE90"   # Verde Claro
}

# Función para obtener la hora correspondiente a un número entero en intervalos de 15 minutos
def getHoraxIntervalo(numero):
    """
    Devuelve la hora en formato HH:MM correspondiente al número entero dado,
    basado en intervalos de 15 minutos desde las 6:00 hasta las 24:00.
    
    :param numero: int, el número entero que representa el intervalo.
    :return: str, la hora en formato HH:MM.
    """
    if numero < 1 or numero > 73:
        raise ValueError("El número debe estar entre 1 y 73.")
    
    base_hora = 6
    intervalos_por_hora = 4  # 4 intervalos de 15 minutos por hora
    minutos_por_intervalo = 15
    
    intervalo = numero - 1
    hora = base_hora + intervalo // intervalos_por_hora
    minutos = (intervalo % intervalos_por_hora) * minutos_por_intervalo
    
    return f"{hora:02}:{minutos:02}"

# Función para llenar las entradas en el Excel
def llenar_entradas(ws, Entradas):
    """
    Escribe los datos del arreglo 'Entradas' en las columnas correspondientes de la hoja de trabajo (ws).
    
    :param ws: La hoja de trabajo donde se escribirán los datos de Entradas.
    :param Entradas: Matriz con las entradas a escribir.
    """
    for i in range(1, 73):  # Recorrer las horas de entrada (1 a 72)
        hora = getHoraxIntervalo(i)  # Obtener la hora en formato HH:MM
        
        # Colocar la hora en la columna A
        celda_hora = f"A{i+2}"  # La columna A, desde la fila 3
        ws[celda_hora] = hora

        for j in range(5):  # Hay hasta 5 columnas por fila
            if Entradas[i][j] != ["", ""]:  # Si la entrada no está vacía
                nombre, tipo_personal = Entradas[i][j]
                
                # Definir la celda correspondiente en la hoja
                celda_columna = get_column_letter(j + 2)  # Desde la columna B (que es la columna 2)
                celda_fila = i + 2  # Fila empieza en 3 porque fila 1 y 2 son los encabezados
                celda = f"{celda_columna}{celda_fila}"
                
                # Escribir el nombre en la celda
                ws[celda] = nombre
                
                # Aplicar color según el tipo de personal
                if tipo_personal in roles_colores:
                    fill = PatternFill(start_color=roles_colores[tipo_personal], end_color=roles_colores[tipo_personal], fill_type='solid')
                    ws[celda].fill = fill
                
                # Aplicar bordes a la celda
                border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )
                ws[celda].border = border

# Función para crear el archivo Excel con el formato de la tabla de horarios
def crear_tabla_excel(Entradas):
    """
    Crea un archivo Excel con el formato de tabla para los horarios del personal,
    llenando los datos desde la matriz 'Entradas'.
    
    :param Entradas: Matriz con las entradas del personal.
    """
    # Crear el libro y la hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Horarios"

    # Configurar encabezados
    ws["A1"] = "Hora"
    ws["B1"] = "Caja 1"
    ws["C1"] = "Caja 2"
    ws["D1"] = "Caja 3"
    ws["E1"] = "Caja 4"
    ws["F1"] = "Caja 5"

    # Llenar las entradas en la hoja
    llenar_entradas(ws, Entradas)

    # Guardar el archivo Excel
    wb.save("horario_personal.xlsx")
    print("Archivo 'horario_personal.xlsx' creado correctamente.")

# Ejemplo para llenar la matriz 'Entradas' con algunos datos
Entradas = [[["", ""] for _ in range(5)] for _ in range(74)]  # Matriz vacía inicial

# Rellenamos algunos datos en Entradas como ejemplo
Entradas[2][0] = ["Maria", "Cajer@"]
Entradas[2][1] = ["Jose", "RS"]
Entradas[10][2] = ["Ana", "Self Checkout"]
Entradas[15][3] = ["Luis", "Ecommerce"]
Entradas[20][4] = ["Pedro", "Supervisor(@)"]

# Llamamos a la función para crear la tabla Excel
crear_tabla_excel(Entradas)
