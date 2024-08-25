import csv
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from collections import defaultdict
import os

def cargar_horarios(archivo):
    horarios = []
    with open(archivo, "r") as archivo:
        lector = csv.reader(archivo)
        for fila in lector:
            horarios.append(fila)
    return horarios

def formatear_hora(hora):
    hora_str = str(hora)
    if len(hora_str) == 1 or len(hora_str) == 2:
        hora_str = hora_str.zfill(2) + ":00"
    elif len(hora_str) == 3:
        hora_str = "0" + hora_str[0] + ":" + hora_str[1:]
    elif len(hora_str) == 4:
        hora_str = hora_str[:2] + ":" + hora_str[2:]
    return hora_str

def obtener_horarios_por_dia(horarios, dia, tipo_personal):
    dia_index = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"].index(dia.lower())
    horarios_dia = defaultdict(list)
    
    for registro in horarios:
        apellidos, nombres, cargo, *dias = registro
        if cargo.lower() == tipo_personal.lower():
            entrada = dias[dia_index * 2]
            salida = dias[dia_index * 2 + 1]
            entrada_formateada = formatear_hora(entrada)
            salida_formateada = formatear_hora(salida)
            if entrada_formateada != "DESCANSO" and salida_formateada != "DESCANSO":
                horarios_dia[(nombres, apellidos)] = (entrada_formateada, salida_formateada)
    
    # Ordenar por hora de entrada
    horarios_dia_ordenados = dict(sorted(horarios_dia.items(), key=lambda x: x[1][0]))
    
    return horarios_dia_ordenados

def generar_intervalos():
    intervalos = []
    for hora in range(6, 24):
        for minuto in [0, 15, 30, 45]:
            intervalos.append(f"{str(hora).zfill(2)}:{str(minuto).zfill(2)}")
    return intervalos

def crear_hoja(ws, intervalos, horarios_dia, tipo_personal):
    ws.append(["Intervalos"] + [f"{nombres} {apellidos}" for nombres, apellidos in horarios_dia.keys()])
    
    for i, intervalo in enumerate(intervalos, start=2):
        ws[f"A{i}"] = intervalo
    
    for col, (nombres, apellidos) in enumerate(horarios_dia.keys(), start=2):
        entrada, salida = horarios_dia[(nombres, apellidos)]
        for i, intervalo in enumerate(intervalos, start=2):
            if entrada <= intervalo < salida:
                ws.cell(row=i, column=col).value = "EN TURNO"
                ws.cell(row=i, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    chart = BarChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=len(intervalos) + 1, max_col=ws.max_column)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, f"E{len(intervalos) + 3}")

def generar_excel(horarios, dia, tipo_personal):
    intervalos = generar_intervalos()
    wb = openpyxl.Workbook()
    
    if tipo_personal == "Todos":
        tipos_personal = ["Asistente de Self Checkout", "Representante de Servicio", "Cajer@", "Ecommerce"]
        for tipo in tipos_personal:
            ws = wb.create_sheet(title=tipo)
            horarios_dia = obtener_horarios_por_dia(horarios, dia, tipo)
            crear_hoja(ws, intervalos, horarios_dia, tipo)
    else:
        ws = wb.active
        ws.title = tipo_personal
        horarios_dia = obtener_horarios_por_dia(horarios, dia, tipo_personal)
        crear_hoja(ws, intervalos, horarios_dia, tipo_personal)
    
    archivo_salida = f"Horario_{tipo_personal}_{dia}.xlsx"
    wb.save(archivo_salida)
    print(f"Archivo {archivo_salida} generado exitosamente.")

def main():
    archivo_horarios = "horario.txt"
    horarios = cargar_horarios(archivo_horarios)
    
    print("\nSeleccionar Día de la Semana:")
    dias_semana = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]
    for i, dia in enumerate(dias_semana, 1):
        print(f"{i}. {dia.capitalize()}")
    dia_opcion = int(input("Ingrese el número de opción: ")) - 1
    dia = dias_semana[dia_opcion]
    
    print("\nSeleccionar Tipo de Personal:")
    tipos_personal = ["Asistente de Self Checkout", "Representante de Servicio", "Cajer@", "Ecommerce", "Todos"]
    for i, tipo in enumerate(tipos_personal, 1):
        print(f"{i}. {tipo}")
    tipo_opcion = int(input("Ingrese el número de opción: ")) - 1
    tipo_personal = tipos_personal[tipo_opcion]
    
    generar_excel(horarios, dia, tipo_personal)

if __name__ == "__main__":
    main()
