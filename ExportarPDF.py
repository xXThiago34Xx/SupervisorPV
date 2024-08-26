import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from PIL import Image
import excel2img

def capturar_hoja_como_imagen(hoja, archivo_imagen):
    excel2img.export_img(hoja, archivo_imagen)

def exportar_imagen_a_pdf(imagen_path, pdf_path, orientacion='portrait'):
    img = Image.open(imagen_path)
    c = canvas.Canvas(pdf_path, pagesize=A4 if orientacion == 'portrait' else landscape(A4))
    width, height = (A4 if orientacion == 'portrait' else landscape(A4))
    img_width, img_height = img.size
    # Ajustar la imagen al tamaño de la página
    img_ratio = img_width / img_height
    page_ratio = width / height
    if img_ratio > page_ratio:
        new_width = width
        new_height = width / img_ratio
    else:
        new_height = height
        new_width = height * img_ratio
    x = (width - new_width) / 2
    y = (height - new_height) / 2
    c.drawImage(imagen_path, x, y, new_width, new_height)
    c.showPage()
    c.save()

def procesar_excel_y_exportar_pdf(input_file, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    wb = load_workbook(input_file)
    for sheet_name in wb.sheetnames:
        orientacion = 'landscape' if 'cajer@' in sheet_name.lower() else 'portrait'
        imagen_filename = f"{sheet_name}.png"
        imagen_path = os.path.join(output_dir, imagen_filename)
        pdf_filename = f"{sheet_name}.pdf"
        pdf_path = os.path.join(output_dir, pdf_filename)

        # Exportar hoja como imagen
        capturar_hoja_como_imagen(input_file, imagen_path)

        # Exportar imagen a PDF
        exportar_imagen_a_pdf(imagen_path, pdf_path, orientacion)

def unir_pdfs(pdf_files, output_pdf):
    from PyPDF2 import PdfMerger
    merger = PdfMerger()
    for pdf in pdf_files:
        merger.append(pdf)
    
    with open(output_pdf, 'wb') as output_file:
        merger.write(output_file)

def main():
    input_file = input("Ingrese la ruta del archivo Excel generado por newGrafic.py: ").strip()
    output_dir = "pdf_exportados"
    
    procesar_excel_y_exportar_pdf(input_file, output_dir)
    
    # Unir todos los PDFs generados
    pdf_files = [os.path.join(output_dir, pdf) for pdf in os.listdir(output_dir) if pdf.endswith('.pdf')]
    output_pdf = os.path.join(output_dir, "documento_final.pdf")
    unir_pdfs(pdf_files, output_pdf)
    
    print(f"Todos los PDFs han sido exportados y unidos en: {output_pdf}")

if __name__ == "__main__":
    main()
